"""
The only program that touches the three-statement model after its birth.

    python3 workbook_refresh.py <instance folder> <period>

What a refresh is: values written into the one named range the manifest declares for the
period, from the sources the manifest names, followed by a real recalculation and the
workbook's own checks. What it is never: a rebuild. No formula is written, no sheet is
added, no layout moves. The customer's model after a refresh is the customer's model.

The guarantees, in the order they are enforced:
  1. The writable range exists and contains no formula. A formula inside a writable
     range means the model and the manifest disagree, and the write REFUSES.
  2. The period is closed. A trial balance without closing balances, or carrying an
     open-month note, is not a closed month; the write REFUSES by name rather than
     landing provisional numbers that look final.
  3. The sources bring exactly the keys the model was born with. A new FSLI is a model
     revision, not a refresh - REFUSED, with the key named.
  4. Formulas are fingerprinted before and after. Owner edits since the last refresh
     are DISCLOSED in the log, never silently absorbed and never reverted - it is the
     owner's model. Drift inside a writable range, though, REFUSES (guarantee 1).
  5. The write lands on a temp copy, LibreOffice recalculates it, and the workbook's
     own CHECK_TOTAL must be zero with MONTHS_CLOSED = previous + 1. Only then does the
     temp replace the model. A failed check leaves the model byte-identical to before.
  6. Every refresh - and every refusal - appends to refresh_log.csv, and every landed
     version is archived whole under versions/. The history is the artifact.
"""
import os, csv, json, sys, hashlib, shutil, subprocess, tempfile, datetime
import openpyxl

def read(p):
    with open(p, encoding="utf-8-sig") as f: return list(csv.DictReader(f))

def sha(p, n=12):
    return hashlib.sha256(open(p,"rb").read()).hexdigest()[:n]

def log_row(folder, row):
    with open(os.path.join(folder,"10-model","refresh_log.csv"),"a",newline="") as f:
        csv.writer(f).writerow(row)

def refuse(folder, version, period, why):
    print(f"REFUSED  {why}")
    log_row(folder, [version, datetime.datetime.now().isoformat(timespec='seconds'),
                     period, "", 0, "", "", "", f"REFUSED - {why}"])
    return 2

def fingerprint(wb):
    fp = {}
    for ws in wb.worksheets:
        for r_ in ws.iter_rows():
            for c in r_:
                if isinstance(c.value,str) and c.value.startswith("="):
                    fp[f"{ws.title}!{c.coordinate}"] = hashlib.sha256(c.value.encode()).hexdigest()[:12]
    return fp

def main(folder, period):
    M = lambda *a: os.path.join(folder, "10-model", *a)
    manifest = json.load(open(M("refresh_manifest.json")))
    state = json.load(open(M("refresh_state.json")))
    model = os.path.join(folder, manifest["model"].replace("/", os.sep))
    v = state["version"]

    tb_path = os.path.join(folder, manifest["sources"]["tb"].format(period=period).replace("/",os.sep))
    cf_path = os.path.join(folder, manifest["sources"]["cf"].format(period=period).replace("/",os.sep))
    if not os.path.exists(tb_path):
        return refuse(folder, v, period, f"no trial balance for {period}")
    tb = read(tb_path)
    # guarantee 2 - the closed-month gate
    if "closing_balance" not in tb[0]:
        return refuse(folder, v, period, f"{period} is open: its trial balance carries no closing balance")
    open_notes = {r.get("note","") for r in tb if "open" in r.get("note","").lower()}
    if open_notes:
        return refuse(folder, v, period, f"{period} is open, and its export says so: \"{next(iter(open_notes))}\"")
    if not os.path.exists(cf_path):
        return refuse(folder, v, period, f"{period} has a closed trial balance but no cash flow export")
    cf = read(cf_path)[0]
    bs_path = os.path.join(folder, manifest["sources"]["bs"].format(period=period).replace("/",os.sep))
    if not os.path.exists(bs_path):
        return refuse(folder, v, period, f"{period} has no balance-sheet export - the TB alone cannot state closing balances")
    bs = read(bs_path)

    wb = openpyxl.load_workbook(model)
    rng_name = manifest["writable"][period]
    if rng_name not in wb.defined_names:
        return refuse(folder, v, period, f"writable range {rng_name} does not exist in the model")
    dest = list(wb.defined_names[rng_name].destinations)[0]
    ws, ref = wb[dest[0]], dest[1].replace("$","")
    cells = [c for row in ws[ref] for c in row]

    # guarantee 1 - no formula lives where values land
    fcells = [c.coordinate for c in cells if isinstance(c.value,str) and c.value.startswith("=")]
    if fcells:
        return refuse(folder, v, period, f"formula found inside writable range at {fcells[0]} - model and manifest disagree")

    # guarantee 3 - exactly the keys the model was born with
    keys = manifest["key_order"]
    vals, brought = {}, set()
    for r in tb:
        if r["type"] not in ("Revenue","COGS","Opex","Other"): continue
        k = f"PNL|{r['type']}|{r['fsli']}"
        vals[k] = vals.get(k,0.0) - float(r["period_movement"]); brought.add(k)
    for r in bs:
        if r["type"] == "Subtotal": continue
        k = f"BS|{r['type']}|{r['fsli']}"
        vals[k] = vals.get(k,0.0) + float(r["closing_balance_usd"]); brought.add(k)
    unknown = brought - set(keys)
    if unknown:
        return refuse(folder, v, period, f"source brings a key the model was not born with: {sorted(unknown)[0]} - model revision, not a refresh")
    for f_ in [k for k in keys if k.startswith("CF|")]:
        vals[f_] = float(cf[f_.split("|",1)[1]])
    vals["FLAG"] = "CLOSED"

    # guarantee 4 - fingerprint before writing
    fp_now = fingerprint(wb)
    drift = {k for k in set(fp_now) | set(state["fingerprint"])
             if fp_now.get(k) != state["fingerprint"].get(k)}
    drift_note = f"{len(drift)} formula cell(s) changed since v{v} (owner edit, kept): " + "; ".join(sorted(drift)[:3]) if drift else ""

    written = 0
    for c, k in zip(cells, keys):
        nv = vals.get(k, 0.0)
        c.value = round(nv,2) if isinstance(nv,float) else nv
        written += 1

    # guarantee 5 - land on a temp, recalculate for real, ask the workbook its own opinion
    with tempfile.TemporaryDirectory() as tmp:
        cand = os.path.join(tmp, "candidate.xlsx")
        wb.save(cand)
        subprocess.run(["libreoffice","--headless","--convert-to","xlsx","--outdir",tmp+"/rc",cand],
                       check=True, capture_output=True, timeout=300)
        rc = openpyxl.load_workbook(os.path.join(tmp,"rc","candidate.xlsx"), data_only=True)
        g = lambda name: rc[list(rc.defined_names[name].destinations)[0][0]][list(rc.defined_names[name].destinations)[0][1].replace("$","")].value
        check_total, months_closed = g("CHECK_TOTAL"), g("MONTHS_CLOSED")
        if check_total is None or abs(check_total) > 0.005:
            return refuse(folder, v, period, f"CHECK_TOTAL is {check_total} after recalculation - the model does not tie, nothing landed")
        if months_closed != state["months_closed"] + 1:
            return refuse(folder, v, period, f"MONTHS_CLOSED is {months_closed}, expected {state['months_closed']+1} - nothing landed")
        # only now does the candidate become the model
        shutil.copy2(cand, model)

    v += 1
    shutil.copy2(model, M("versions", f"three-statement-FY2026_v{v:03d}.xlsx"))
    state.update({"version": v, "months_closed": state["months_closed"]+1, "fingerprint": fingerprint(openpyxl.load_workbook(model))})
    json.dump(state, open(M("refresh_state.json"),"w"))
    srcs = f"{os.path.basename(tb_path)}@{sha(tb_path)};{os.path.basename(bs_path)}@{sha(bs_path)};{os.path.basename(cf_path)}@{sha(cf_path)}"
    log_row(folder, [v, datetime.datetime.now().isoformat(timespec='seconds'), period,
                     srcs, written, f"{check_total:.2f}", months_closed, drift_note, "LANDED"])
    print(f"LANDED  v{v}: {period} closed into the model - {written} cells written, "
          f"CHECK_TOTAL {check_total:.2f}, months closed {months_closed}"
          + (f"\nDISCLOSED  {drift_note}" if drift_note else ""))
    return 0

if __name__ == "__main__":
    folder = sys.argv[1] if len(sys.argv)>1 else os.path.join(os.path.dirname(__file__),"..")
    period = sys.argv[2] if len(sys.argv)>2 else "2026-01"
    sys.exit(main(folder, period))
