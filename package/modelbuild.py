"""
The three-statement model, built ONCE.

Every other workbook in this instance is rebuilt by its engine each period. This one is
the opposite experiment, and the point of G1: a long-lived artifact that is constructed
a single time and then only ever REFRESHED - values written into declared ranges,
formulas never regenerated, version history kept. After this script runs, the only
program that may touch the file is workbook_refresh.py, and the manifest it emits here
is the contract between them: which cells are writable, from which sources, and which
checks must hold after every write.

Layout: Data_Plan (plan by P&L line, written once here), Data_Actuals (written month by
month by refresh, empty at birth), P&L / Balance sheet / Cash flow (formulas only - a
closed month reads actuals, an open month reads plan or blank), Checks (the workbook's
own tie-outs; refresh refuses any write after which these are not zero).

Signs: P&L stored credit-positive (revenue +, costs -), so a column sums to net result.
Balance-sheet rows carry the trial balance's own signs; presentation flips them.
"""
import os, csv, json, sys, hashlib, shutil
import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

MONTHS = [f"2026-{m:02d}" for m in range(1, 13)]

PNL_TYPES = ("Revenue", "COGS", "Opex", "Other")
CF_FIELDS = ["net_result","stock_based_compensation","depreciation_and_amortization",
             "credit_loss_provision","working_capital_movement","cash_from_operations",
             "capital_expenditure","financing","free_cash_flow","opening_cash",
             "net_cash_movement","closing_cash","unexplained","da_per_pnl","ecl_per_pnl"]

def read(p):
    with open(p, encoding="utf-8-sig") as f: return list(csv.DictReader(f))

def build(folder):
    P = lambda *a: os.path.join(folder, *a)
    tb = read(P("03-actuals","FY2026","trial_balance_2026-01.csv"))
    plan = read(P("02-budget","FY2026","plan_fy26_pnl_monthly.csv"))
    # BS keys are born from the balance-sheet export, not the TB: the FY2026 trial
    # balance carries YTD movements with no opening balances, which the first refresh
    # attempt refused on. The export balances to the cent once its Subtotal rows go.
    bsrows = read(P("03-actuals","FY2026","balance_sheet_2026-01.csv"))

    # ---- the model's row keys, fixed at birth. A source that later brings a key not in
    # this list is a model revision, not a refresh, and refresh will refuse it.
    pnl_keys, bs_keys, seen = [], [], set()
    for r in tb:
        if r["type"] in PNL_TYPES:
            k = ("PNL", r["type"], r["fsli"])
            if k not in seen: seen.add(k); pnl_keys.append(k)
    for r in bsrows:
        if r["type"] == "Subtotal": continue
        k = ("BS", r["type"], r["fsli"])
        if k not in seen: seen.add(k); bs_keys.append(k)
    order = {"Revenue":0,"COGS":1,"Opex":2,"Other":3}
    pnl_keys.sort(key=lambda k:(order[k[1]], k[2]))
    bs_keys.sort(key=lambda k:(("Asset","Liability","Equity").index(k[1]), k[2]))

    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # ---- Data_Plan: P&L keys x months, credit-positive, values written once, here.
    dp = wb.create_sheet("Data_Plan")
    dp["A1"] = "key"
    for j,m in enumerate(MONTHS): dp.cell(row=1,column=2+j,value=m)
    pmap = {}
    for r in plan:
        k = ("PNL", None, r["fsli"]); pmap.setdefault((r["fsli"], r["period"]), 0.0)
        pmap[(r["fsli"], r["period"])] += -float(r["amount_usd"])
    for i,k in enumerate(pnl_keys):
        dp.cell(row=2+i,column=1,value="|".join(k))
        for j,m in enumerate(MONTHS):
            dp.cell(row=2+i,column=2+j,value=round(pmap.get((k[2],m),0.0),2))

    # ---- Data_Actuals: same P&L keys, then BS keys, then CF fields, then the flag row.
    da = wb.create_sheet("Data_Actuals")
    da["A1"] = "key"
    for j,m in enumerate(MONTHS): da.cell(row=1,column=2+j,value=m)
    keys = [ "|".join(k) for k in pnl_keys ] + [ "|".join(k) for k in bs_keys ] \
         + [ f"CF|{f}" for f in CF_FIELDS ] + ["FLAG"]
    for i,k in enumerate(keys): da.cell(row=2+i,column=1,value=k)
    n_pnl, n_bs, n_cf = len(pnl_keys), len(bs_keys), len(CF_FIELDS)
    r_pnl, r_bs, r_cf = 2, 2+n_pnl, 2+n_pnl+n_bs
    r_flag = 2+n_pnl+n_bs+n_cf
    col = lambda j: get_column_letter(2+j)
    flag = lambda j: f"Data_Actuals!${col(j)}${r_flag}"
    cfrow = lambda f: r_cf + CF_FIELDS.index(f)

    # ---- P&L: closed months read actuals, open months read plan.
    pl = wb.create_sheet("P&L"); pl["A1"]="P&L - FY2026"; pl["A2"]="basis"
    for j,m in enumerate(MONTHS):
        pl.cell(row=1,column=2+j,value=m)
        pl.cell(row=2,column=2+j,value=f'=IF({flag(j)}="CLOSED","ACTUAL","PLAN")')
    pl.cell(row=1,column=14,value="FY")
    row = 3; pl_rows = {}
    for i,k in enumerate(pnl_keys):
        pl.cell(row=row,column=1,value=k[2]); pl_rows[k]=row
        for j in range(12):
            c = f"{col(j)}{r_pnl+i}"
            pl.cell(row=row,column=2+j,
                    value=f'=IF({flag(j)}="CLOSED",Data_Actuals!{c},Data_Plan!{c})')
        pl.cell(row=row,column=14,value=f"=SUM(B{row}:M{row})")
        row += 1
    def subtotal(label, member_rows):
        nonlocal row
        pl.cell(row=row,column=1,value=label)
        for j in range(12):
            L = get_column_letter(2+j)
            pl.cell(row=row,column=2+j,value="=" + "+".join(f"{L}{r}" for r in member_rows))
        pl.cell(row=row,column=14,value=f"=SUM(B{row}:M{row})")
        row += 1; return row-1
    gp = subtotal("Gross profit", [pl_rows[k] for k in pnl_keys if k[1] in ("Revenue","COGS")])
    op = subtotal("Operating result", [gp] + [pl_rows[k] for k in pnl_keys if k[1]=="Opex"])
    ni = subtotal("Net result", [op] + [pl_rows[k] for k in pnl_keys if k[1]=="Other"])

    # ---- Balance sheet: closed months only; liabilities and equity flipped to display sign.
    bs = wb.create_sheet("Balance sheet"); bs["A1"]="Balance sheet - FY2026 (closed months)"
    for j,m in enumerate(MONTHS): bs.cell(row=1,column=2+j,value=m)
    row = 2; bs_rows = {}
    def bs_line(label, formula_fn):
        nonlocal row
        bs.cell(row=row,column=1,value=label)
        for j in range(12):
            bs.cell(row=row,column=2+j,value=formula_fn(j))
        row += 1; return row-1
    a_rows = []
    for i,k in enumerate(bs_keys):
        if k[1]!="Asset": continue
        r = bs_line(k[2], lambda j,i=i: f'=IF({flag(j)}="CLOSED",Data_Actuals!{col(j)}{r_bs+i},"")')
        bs_rows[k]=r; a_rows.append(r)
    ta = bs_line("TOTAL ASSETS", lambda j: f'=IF({flag(j)}="CLOSED",'+ "+".join(f"{get_column_letter(2+j)}{r}" for r in a_rows) + ',"")')
    l_rows = []
    for i,k in enumerate(bs_keys):
        if k[1]!="Liability": continue
        r = bs_line(k[2], lambda j,i=i: f'=IF({flag(j)}="CLOSED",-Data_Actuals!{col(j)}{r_bs+i},"")')
        l_rows.append(r)
    tl = bs_line("TOTAL LIABILITIES", lambda j: f'=IF({flag(j)}="CLOSED",'+ "+".join(f"{get_column_letter(2+j)}{r}" for r in l_rows) + ',"")')
    e_rows = []
    for i,k in enumerate(bs_keys):
        if k[1]!="Equity": continue
        r = bs_line(k[2], lambda j,i=i: f'=IF({flag(j)}="CLOSED",-Data_Actuals!{col(j)}{r_bs+i},"")')
        e_rows.append(r)
    # no retained-earnings plug: the export's equity already carries current-year results
    te = bs_line("TOTAL EQUITY", lambda j: f'=IF({flag(j)}="CLOSED",'+ "+".join(f"{get_column_letter(2+j)}{r}" for r in e_rows) + ',"")')
    cash_key = next(k for k in bs_keys if k[1]=="Asset" and "Cash and cash" in k[2])
    cash_row = bs_rows[cash_key]

    # ---- Cash flow: straight presentation of the CF block, closed months only.
    cf = wb.create_sheet("Cash flow"); cf["A1"]="Cash flow - FY2026 (closed months, indirect)"
    for j,m in enumerate(MONTHS): cf.cell(row=1,column=2+j,value=m)
    for i,f_ in enumerate(CF_FIELDS):
        cf.cell(row=2+i,column=1,value=f_)
        for j in range(12):
            cf.cell(row=2+i,column=2+j,value=f'=IF({flag(j)}="CLOSED",Data_Actuals!{col(j)}{r_cf+i},"")')

    # ---- Checks: the workbook's own tie-outs, all zero or the refresh does not land.
    ck = wb.create_sheet("Checks"); ck["A1"]="check"; ck["N1"]="TOTAL"
    for j,m in enumerate(MONTHS): ck.cell(row=1,column=2+j,value=m)
    defs = [
        ("BS balances: assets - liabilities - equity",
         lambda j: f'=IF({flag(j)}="CLOSED",ROUND(\'Balance sheet\'!{get_column_letter(2+j)}{ta}-\'Balance sheet\'!{get_column_letter(2+j)}{tl}-\'Balance sheet\'!{get_column_letter(2+j)}{te},2),0)'),
        ("CF closing cash = BS cash",
         lambda j: f'=IF({flag(j)}="CLOSED",ROUND(\'Cash flow\'!{get_column_letter(2+j)}{2+CF_FIELDS.index("closing_cash")}-\'Balance sheet\'!{get_column_letter(2+j)}{cash_row},2),0)'),
        ("CF net result = P&L net result",
         lambda j: f'=IF({flag(j)}="CLOSED",ROUND(\'Cash flow\'!{get_column_letter(2+j)}{2+CF_FIELDS.index("net_result")}-\'P&L\'!{get_column_letter(2+j)}{ni},2),0)'),
        ("CF opening + movement = closing",
         lambda j: f'=IF({flag(j)}="CLOSED",ROUND(\'Cash flow\'!{get_column_letter(2+j)}{2+CF_FIELDS.index("opening_cash")}+\'Cash flow\'!{get_column_letter(2+j)}{2+CF_FIELDS.index("net_cash_movement")}-\'Cash flow\'!{get_column_letter(2+j)}{2+CF_FIELDS.index("closing_cash")},2),0)'),
        ("CF unexplained is zero",
         lambda j: f'=IF({flag(j)}="CLOSED",ROUND(\'Cash flow\'!{get_column_letter(2+j)}{2+CF_FIELDS.index("unexplained")},2),0)'),
    ]
    for i,(label,fn) in enumerate(defs):
        ck.cell(row=2+i,column=1,value=label)
        for j in range(12): ck.cell(row=2+i,column=2+j,value=fn(j))
        ck.cell(row=2+i,column=14,value=f"=ROUND(SUMPRODUCT(ABS(B{2+i}:M{2+i})),2)")
    tot_row = 2+len(defs)
    ck.cell(row=tot_row,column=1,value="CHECK_TOTAL (must be 0)")
    ck.cell(row=tot_row,column=14,value=f"=ROUND(SUM(N2:N{1+len(defs)}),2)")
    mc_row = tot_row+1
    ck.cell(row=mc_row,column=1,value="MONTHS_CLOSED")
    ck.cell(row=mc_row,column=14,value=f'=COUNTIF(Data_Actuals!$B${r_flag}:$M${r_flag},"CLOSED")')

    # ---- named ranges: the refresh contract.
    last = r_flag
    for j,m in enumerate(MONTHS):
        wb.defined_names.add(DefinedName(f"ACT_{m.replace('-','_')}",
            attr_text=f"Data_Actuals!${col(j)}$2:${col(j)}${last}"))
    wb.defined_names.add(DefinedName("CHECK_TOTAL", attr_text=f"Checks!$N${tot_row}"))
    wb.defined_names.add(DefinedName("MONTHS_CLOSED", attr_text=f"Checks!$N${mc_row}"))

    out_dir = P("10-model"); os.makedirs(out_dir, exist_ok=True)
    model = os.path.join(out_dir, "three-statement-FY2026.xlsx")
    wb.save(model)

    # ---- manifest, fingerprint, version 1, empty log.
    manifest = {
        "model": "10-model/three-statement-FY2026.xlsx",
        "writable": {m: f"ACT_{m.replace('-','_')}" for m in MONTHS},
        "key_order": keys,
        "signs": {"PNL": "-period_movement (credit-positive)", "BS": "closing_balance as booked", "CF": "as exported"},
        "sources": {"tb": "03-actuals/FY2026/trial_balance_{period}.csv",
                    "bs": "03-actuals/FY2026/balance_sheet_{period}.csv",
                    "cf": "03-actuals/FY2026/cash_flow_{period}.csv"},
        "closed_month_gate": "tb must carry closing_balance and no open-month note",
        "checks": "CHECK_TOTAL == 0 and MONTHS_CLOSED == previous + 1 after recalculation",
    }
    json.dump(manifest, open(os.path.join(out_dir,"refresh_manifest.json"),"w"), indent=1)

    fp = {}
    for ws in wb.worksheets:
        for r_ in ws.iter_rows():
            for c in r_:
                if isinstance(c.value,str) and c.value.startswith("="):
                    fp[f"{ws.title}!{c.coordinate}"] = hashlib.sha256(c.value.encode()).hexdigest()[:12]
    state = {"version": 1, "months_closed": 0, "formula_cells": len(fp), "fingerprint": fp}
    json.dump(state, open(os.path.join(out_dir,"refresh_state.json"),"w"))
    os.makedirs(os.path.join(out_dir,"versions"), exist_ok=True)
    shutil.copy2(model, os.path.join(out_dir,"versions","three-statement-FY2026_v001.xlsx"))
    with open(os.path.join(out_dir,"refresh_log.csv"),"w",newline="") as f:
        csv.writer(f).writerow(["version","timestamp","period","sources","cells_written",
                                "check_total","months_closed","formula_drift","verdict"])
    print(f"built: {len(keys)} data rows, {len(fp)} formula cells fingerprinted, v001 archived")
    print("from here on, only workbook_refresh.py touches this file")

if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv)>1 else os.path.join(os.path.dirname(__file__),".."))
