"""
Does the write-up still agree with the artifact it describes?

Nothing else in this instance reads prose. packverify.py asserts the workbook against the
instance CSVs; validate.py asserts the instance against itself. Both pass while a document
quoting those workbooks goes stale, which is exactly what happened between 28 Aug and 1 Sep:
the pack was rebuilt, the write-up was not, and the only figures a reader could see were the
wrong ones.

v1 checked one write-up and printed a note naming the ones it did not check. That note listed
25 documents, and a list that never shrinks is a list nobody reads. v2 replaces the note with
a rule:

    every write-up that quotes figures carries an explicit classification, and an
    unclassified figure-heavy write-up FAILS the build.

Four classifications, because write-ups make four different kinds of claim:

    current       describes a shipped artifact. Every quoted figure must exist in the
                  RECALCULATED artifact - the v1 check, unchanged. A current doc from
                  which no figures can be extracted also FAILS: a check that can pass
                  on an empty set is not a check.
    record        a dated account of a run, a fix, or an audit. Its figures were true
                  when it was written and are kept because the failures changed the
                  design. Not gated against today's artifact - gating history against
                  the present would force the record to lie.
    illustrative  charters, contracts, specs, playbooks, the registry. Their figures
                  are worked examples or rulings, not claims about a shipped artifact.
    unverifiable  describes an artifact that is not published in this repository. The
                  honest verdict is a named refusal, not silence: the register says
                  which artifact would be needed.

The classifications are themselves the point. "Not checked" was silence; "record, 25 Aug,
superseded by the shipped pack" is a statement someone can dispute.

One more instrument, because the recurring failure shape in the run log is "a number right
in one place and wrong in another, with nothing putting the two side by side": the register
ends with a SIDE-BY-SIDE section listing every named metric quoted with different values in
different write-ups, with doc and line for each. It is reported, not scored - across dated
records a moving figure is usually the design working - except between two `current` docs,
where a disagreement on the same metric and period is a FAIL.

The workbooks carry no cached values, so this reads RECALCULATED values, the same way
packverify does. A checker that trusts what is typed in the file it is checking has checked
nothing.

    python3 docverify.py [repo root]

Writes outputs/docverify-register.md. Exempt a line that legitimately quotes a figure from
outside the artifact with a trailing marker:  <!-- docverify: external -->
"""
import os, re, sys, subprocess, tempfile
from openpyxl import load_workbook

ROOT = os.path.abspath(sys.argv[1] if len(sys.argv) > 1 else
                       os.path.join(os.path.dirname(__file__), ".."))
REGISTER = os.path.join(ROOT, "outputs", "docverify-register.md")

# ---------------------------------------------------------------------------------------
# The classification table. Every write-up that quotes figures appears here, on purpose,
# so that adding a figure-heavy doc without deciding what kind of claim it makes is a
# failure rather than an omission. Paths are repo-relative.
# ---------------------------------------------------------------------------------------

CURRENT = {
    "outputs/arcline-pack-and-lbe.md": [
        "outputs/FY2026-01-management-pack.xlsx",
        "outputs/LBE_Q1_2026_M1.xlsx",
    ],
}

RECORD = {
    "outputs/management-reporting-pack.md":  "25 Aug build record; superseded by the shipped 9-tab pack",
    "outputs/pack-rebuilt.md":               "18 Aug rebuild record",
    "outputs/entry-triage-and-reforecast.md":"24 Aug run record",
    "runs/first-agent-run.md":               "run record",
    "runs/full-sweep-findings.md":           "run record",
    "runs/ingestion-at-volume.md":           "run record",
    "runs/run-log.md":                       "the run log; every figure is dated by its block",
    "runs/scorecard.md":                     "scorecard snapshot at publication",
    "runs/sealed-run-01.md":                 "sealed-run record",
    "runs/trigger-table.md":                 "trigger arming record",
    "runs/arcline/01-test-instance.md":      "instance build record",
    "runs/arcline/02-run-01.md":             "run record",
    "runs/arcline/03-instrument-fixes.md":   "fix record",
    "runs/arcline/04-run-03-findings.md":    "run record",
    "correction-loop/first-review-session.md":"review record",
    "correction-loop/reviewer-edits.md":     "review record",
    "correction-loop/loop-verification.md":  "verification record",
    "correction-loop/self-improvement-loop.md":"loop design record",
    "correction-loop/observability.md":      "KPI snapshot at publication",
    "red-team/audit-brief.md":               "audit record",
    "red-team/audit-findings.md":            "audit record",
    "red-team/audit-response.md":            "audit record",
    "what-broke/failure-case.md":            "kept failure",
    "what-broke/friction-log.md":            "kept failures",
    "what-broke/plan-hash-incident.md":      "kept failure",
    "outputs/pipeline-tier3.md":             "4 Sep run record; artifacts live in the private instance",
}

ILLUSTRATIVE = {
    "agents/analyst.md":                     "charter; worked examples",
    "agents/bookkeeper.md":                  "charter; worked examples",
    "agents/chief-of-staff.md":              "charter; worked examples",
    "agents/controller.md":                  "charter; worked examples",
    "agents/copilot-charter.md":             "charter; worked examples",
    "agents/evidence.md":                    "charter; worked examples",
    "agents/forecaster.md":                  "charter; worked examples",
    "agents/operator-instructions.md":       "operating doc; worked examples",
    "contracts/commentary-contract.md":      "contract; worked examples",
    "playbooks/wiring.md":                   "playbook; worked examples",
    "architecture/saas-layer.md":            "architecture; worked examples",
    "architecture/slicing.md":               "architecture; worked examples",
    "architecture/source-vs-output.md":      "architecture; worked examples",
    "install/org-assessment.md":             "install doc; worked examples",
    "data/calculation-methodology.md":       "generator spec; design parameters",
    "data/dataset-build-notes.md":           "generator spec; design parameters",
    "data/simulated-company-spec.md":        "generator spec; design parameters",
    "data/edge-cases/churn-plant.md":        "planted-defect spec; the figures are the plant",
    "semantic-layer/definitions-instance.md":"the registry; the figures are the rulings",
    "semantic-layer/glossary.md":            "glossary; worked examples",
    "semantic-layer/rulings/plan-ruling-r2.md":"ruling; the figures are the ruling",
}

UNVERIFIABLE = {
    "outputs/forecast-model.md":             "the forecast workbook is not published in this repository",
    "outputs/lrp/long-range-plan.md":        "the LRP workbook is not published in this repository",
    "outputs/lrp/planning-cadence-lbe.md":   "describes the cadence around workbooks not published here",
}

# A new, unclassified doc quoting at least this many figures fails the build. Below the
# threshold it is noted, because two incidental figures do not make a write-up a claim.
UNCLASSIFIED_FAIL_AT = 5

MONEY_TOL = 1.0      # the doc rounds to whole units
PCT_TOL = 0.06       # the doc rounds a percentage to one decimal
DEC_TOL = 0.06       # ... and a day-count or month-count the same way

EXEMPT = "docverify: external"
METRIC_WORDS = ("dso", "runway", "days", "months", "pts")

results = []


def chk(name, ok, detail=""):
    results.append((bool(ok), name, detail))
    print(("PASS  " if ok else "FAIL  ") + name + ("  " + str(detail) if detail else ""))


def recalc(path, out):
    """Open the workbook and let the spreadsheet engine compute it, exactly as a reader would."""
    subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", out, path],
                   check=True, capture_output=True, timeout=420)
    return os.path.join(out, os.path.basename(path))


def workbook_values(paths):
    """Every number the workbooks actually compute, in the three shapes a doc quotes them."""
    money, pct, dec = set(), set(), set()
    with tempfile.TemporaryDirectory() as tmp:
        for p in paths:
            wb = load_workbook(recalc(p, tmp), data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for c in row:
                        v = c.value
                        if not isinstance(v, (int, float)) or isinstance(v, bool):
                            continue
                        money.add(round(abs(float(v))))
                        pct.add(round(abs(float(v)) * 100, 1))   # 0.0256 -> 2.6
                        pct.add(round(abs(float(v)), 1))         # already a percentage
                        dec.add(round(abs(float(v)), 1))
    return money, pct, dec


MONEY_RE = re.compile(r'\(?\b\d{1,3}(?:,\d{3})+(?:\.\d+)?\b\)?')
PCT_RE = re.compile(r'\b\d+\.\d+\s*%')
DEC_RE = re.compile(r'(?<![\d,.])\d+\.\d(?![\d%])')


def figures(line):
    """The figures a reader would take from this line as facts about the artifact."""
    out = []
    for m in MONEY_RE.findall(line):
        out.append(("money", abs(float(m.strip("()").replace(",", "")))))
    for m in PCT_RE.findall(line):
        out.append(("pct", abs(float(m.rstrip("% ").strip()))))
    if any(w in line.lower() for w in METRIC_WORDS):
        for m in DEC_RE.findall(line):
            out.append(("dec", abs(float(m))))
    return out


def near(value, pool, tol):
    return any(abs(value - p) <= tol for p in pool)


# ---------------------------------------------------------------------------------------
# Side by side. Named metrics only, deliberately few, because the point is a short list a
# person actually reads. Each pattern captures one value; the period is read off the same
# line when it is there.
# ---------------------------------------------------------------------------------------

METRICS = [
    ("runway (months)",  re.compile(r'runway[^.\n]{0,40}?(\d+(?:\.\d+)?)\s*month', re.I)),
    ("DSO (days)",       re.compile(r'\bDSO\b[^.\n]{0,30}?(\d+(?:\.\d+)?)', re.I)),
    ("cash",             re.compile(r'\bcash\b[^.\n]{0,30}?\b(?:of|at|is|was|ended[^0-9\n]{0,15})\s*\$?([\d]{1,3}(?:,\d{3})+)', re.I)),
    ("ARR",              re.compile(r'\bARR\b[^.\n]{0,30}?\$?([\d]{1,3}(?:,\d{3})+)', re.I)),
    ("revenue",          re.compile(r'\brevenue\b[^.\n]{0,25}?\b(?:of|is|was|at)\s*\$?([\d]{1,3}(?:,\d{3})+)', re.I)),
    ("gross margin (%)", re.compile(r'gross\s+margin[^.\n]{0,30}?(\d+(?:\.\d+)?)\s*%', re.I)),
    ("NRR (%)",          re.compile(r'\bNRR\b[^.\n]{0,25}?(\d+(?:\.\d+)?)\s*%', re.I)),
    ("GRR (%)",          re.compile(r'\bGRR\b[^.\n]{0,25}?(\d+(?:\.\d+)?)\s*%', re.I)),
]

PERIOD_RE = re.compile(
    r'\b(January|February|March|April|May|June|July|August|September|October|November|December'
    r'|Jan|Feb|Mar|Apr|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[-\s]?(?:20)?(\d{2})?\b'
    r'|\b(20\d{2})-(\d{2})(?!-\d)\b|\b(Q[1-4])\b|\b(FY2\d(?:2\d)?)\b')  # 2026-08, not the 2026-08-22 of a ruling date

MONTHS = {m[:3].lower(): i for i, m in enumerate(
    ["January","February","March","April","May","June","July","August",
     "September","October","November","December"], 1)}


def period_of(line):
    m = PERIOD_RE.search(line)
    if not m:
        return "unstated"
    if m.group(1):
        yy = m.group(2) or "26"
        return f"20{yy}-{MONTHS[m.group(1)[:3].lower()]:02d}"
    if m.group(3):
        return f"{m.group(3)}-{m.group(4)}"
    return m.group(5) or m.group(6)


# "net new ARR" is not ARR. A qualifier just before the metric word means the line is
# talking about a movement or a component, and it stays out of the side-by-side.
QUALIFIED = re.compile(r'(net\s+new|new|churn\w*|contract\w*|expan\w*|delta|change\s+in)\s*$', re.I)


def metric_mentions(rel, text):
    out = []
    for n, line in enumerate(text.splitlines(), 1):
        if EXEMPT in line:
            continue
        for name, rx in METRICS:
            for m in rx.finditer(line):
                if QUALIFIED.search(line[:m.start()]):
                    continue
                out.append((name, period_of(line), float(m.group(1).replace(",", "")), rel, n))
    return out


def main():
    all_md, mentions, rows = [], [], []
    for root, _, files in os.walk(ROOT):
        if os.sep + "." in root:
            continue
        for f in files:
            if f.endswith(".md"):
                rel = os.path.relpath(os.path.join(root, f), ROOT).replace(os.sep, "/")
                if rel != "outputs/docverify-register.md":   # its own output
                    all_md.append(rel)

    classified = set(CURRENT) | set(RECORD) | set(ILLUSTRATIVE) | set(UNVERIFIABLE)
    for rel in classified:
        chk(f"{rel} - classified write-up exists", rel in all_md, "listed in docverify but not in the repo") \
            if rel not in all_md else None

    for rel in sorted(all_md):
        text = open(os.path.join(ROOT, rel), encoding="utf-8").read()
        lines = text.splitlines()
        n_figs = sum(len(figures(l)) for l in lines if EXEMPT not in l)
        mentions += metric_mentions(rel, text)

        if rel in CURRENT:
            paths = [os.path.join(ROOT, a) for a in CURRENT[rel]]
            missing = [p for p in paths if not os.path.exists(p)]
            if missing:
                chk(f"{rel} - its artifacts exist", False, missing)
                rows.append((rel, "current", n_figs, "FAIL - artifact missing"))
                continue
            money, pct, dec = workbook_values(paths)
            pools = {"money": (money, MONEY_TOL), "pct": (pct, PCT_TOL), "dec": (dec, DEC_TOL)}
            orphans, quoted = [], 0
            for n, line in enumerate(lines, 1):
                if EXEMPT in line:
                    continue
                for kind, v in figures(line):
                    quoted += 1
                    pool, tol = pools[kind]
                    if not near(v, pool, tol):
                        orphans.append(f"line {n}: {v:,.1f} ({kind})")
            chk(f"{rel} - a current doc quotes at least one figure", quoted > 0,
                "a check that can pass on an empty set is not a check")
            chk(f"{rel} - every quoted figure exists in the artifact", not orphans,
                f"{quoted} figures checked" + (f", {len(orphans)} orphaned: " + "; ".join(orphans[:6])
                                               if orphans else ""))
            rows.append((rel, "current", quoted,
                         "CONTRADICTS - " + "; ".join(orphans[:3]) if orphans else "TIES"))
        elif rel in RECORD:
            if n_figs:
                rows.append((rel, "record", n_figs, RECORD[rel]))
        elif rel in ILLUSTRATIVE:
            if n_figs:
                rows.append((rel, "illustrative", n_figs, ILLUSTRATIVE[rel]))
        elif rel in UNVERIFIABLE:
            rows.append((rel, "unverifiable", n_figs, "REFUSED - " + UNVERIFIABLE[rel]))
        elif n_figs >= UNCLASSIFIED_FAIL_AT:
            chk(f"{rel} - figure-heavy write-up is classified", False,
                f"{n_figs} figures and no entry in docverify.py - classify it or stop quoting figures")
            rows.append((rel, "UNCLASSIFIED", n_figs, "FAIL"))
        elif n_figs:
            rows.append((rel, "(noted)", n_figs, "below threshold; unclassified"))

    # Two current docs disagreeing on the same STATED metric and period is the one
    # cross-doc disagreement with no innocent reading. An unstated period is not gated:
    # in a table the period lives in the header, not on the line, and gating a guess
    # would make the check cry wolf.
    cur = {}
    for name, period, value, rel, n in mentions:
        if rel in CURRENT and period != "unstated":
            cur.setdefault((name, period), set()).add(round(value, 1))
    for (name, period), vals in sorted(cur.items()):
        if len(vals) > 1:
            chk(f"current docs agree on {name} for {period}", False, sorted(vals))

    # The side-by-side listing: every named metric quoted with more than one value.
    side = {}
    for name, period, value, rel, n in mentions:
        side.setdefault((name, period), []).append((round(value, 1), rel, n))
    conflicts = {k: v for k, v in sorted(side.items())
                 if len({x[0] for x in v}) > 1}

    with open(REGISTER, "w", encoding="utf-8") as f:
        f.write("# Docverify register\n\n")
        f.write("*Generated by `outputs/docverify.py`. Do not edit; rerun. No date on purpose: "
                "CI regenerates this file and diffs it against the committed one, so its content "
                "may depend only on the tree.*\n\n")
        f.write("Every write-up in this repository that quotes figures, and what kind of claim "
                "it makes. `current` is verified against the recalculated artifact; `record` is "
                "a dated account kept as history; `illustrative` figures are worked examples or "
                "rulings; `unverifiable` is a named refusal.\n\n")
        f.write("| Write-up | Class | Figures | Verdict / reason |\n|---|---|---:|---|\n")
        for rel, role, n_figs, verdict in rows:
            f.write(f"| {rel} | {role} | {n_figs} | {verdict} |\n")
        f.write("\n## Side by side\n\n")
        f.write("Named metrics quoted with more than one value across write-ups. Reported, not "
                "scored: across dated records a moving figure is usually the design working. "
                "The point is that the two numbers now sit next to each other.\n\n")
        if not conflicts:
            f.write("No named metric is quoted with conflicting values.\n")
        for (name, period), v in conflicts.items():
            f.write(f"**{name}, {period}**\n\n")
            for value, rel, n in sorted(v):
                f.write(f"- {value:,.1f} — {rel}:{n}\n")
            f.write("\n")

    print(f"\nRegister written to {os.path.relpath(REGISTER, ROOT)} "
          f"({len(rows)} write-ups, {len(conflicts)} side-by-side entries)")
    failed = [r for r in results if not r[0]]
    print(f"{len(results) - len(failed)} of {len(results)} checks pass")
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
