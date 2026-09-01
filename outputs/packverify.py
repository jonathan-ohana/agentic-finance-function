"""
Independent verification of the management pack and the LBE.

Independent BY CONSTRUCTION: it recomputes every figure from the instance CSVs and reads the
workbook's RECALCULATED values, never its formulas. A verifier that reads the same formula
the builder wrote proves the builder is self-consistent and nothing else.

Any pack can foot down a column. The failure that reaches a CEO is two tabs describing the
same month with different numbers, so the cross-tab checks are the ones that matter.
"""
import os, sys, csv, subprocess, tempfile, shutil
from collections import defaultdict
from openpyxl import load_workbook

FOLDER = (sys.argv[1] if len(sys.argv) > 1 else
          os.environ.get("ARCLINE_FOLDER", "/home/claude/out/Arcline-Finance"))
PACK = os.path.join(FOLDER, "08-reporting", "FY2026", "FY2026-01-management-pack.xlsx")
LBE = os.path.join(FOLDER, "05-lbe", "FY2026", "LBE_Q1_2026_M1.xlsx")
FLUX = os.path.join(FOLDER, "03-actuals", "FY2026", "variance_analysis_2026-01.xlsx")
TOL = 0.75

results = []


def chk(name, ok, detail=""):
    results.append((bool(ok), name, detail))
    print(("PASS  " if ok else "FAIL  ") + name + ("  " + str(detail) if detail else ""))


def recalc(path, out):
    subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", out, path],
                   check=True, capture_output=True, timeout=420)
    return os.path.join(out, os.path.basename(path))


def read(p):
    with open(p, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def find(ws, text, col=2, upto=400):
    for r in range(1, min(ws.max_row, upto) + 1):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip() == text:
            return r
    return None


def val(ws, text, col, label_col=2):
    r = find(ws, text, label_col)
    return None if r is None else ws.cell(row=r, column=col).value


def num(x):
    return 0.0 if x in (None, "") else float(x)


def main():
    tmp = tempfile.mkdtemp()
    try:
        wb = load_workbook(recalc(PACK, tmp), data_only=True)
        raw = load_workbook(PACK)                    # formulas, for the formula-ness checks
        lwb = load_workbook(recalc(LBE, tmp), data_only=True)

        P = wb["P&L by FSLI"]; R = wb["Revenue and ARR"]; G = wb["Gross margin"]
        O = wb["Opex by owner"]; B = wb["Balance sheet"]; C = wb["Cash"]; M = wb["SaaS metrics"]
        LB = wb["LBE Q1 M1"]; LS = lwb["LBE_Q1_2026_M1"]

        # ---------------------------------------------------- source of truth, recomputed
        F = lambda *a: os.path.join(FOLDER, *a)
        acc = {r["account"]: r for r in read(F("00-company", "chart_of_accounts.csv"))}
        act = {r["account"]: float(r["amount_usd"])
               for r in read(F("03-actuals", "FY2026", "pnl_2026-01.csv")) if r["account"]}
        pln = {r["account"]: float(r["amount_usd"])
               for r in read(F("02-budget", "FY2026", "plan_fy26_pnl_monthly.csv"))
               if r["account"] and r["period"] == "2026-01"}
        by = lambda d, f: sum(v for a, v in d.items() if acc[a]["fsli"] == f)
        rev_a, rev_p = by(act, "Revenue"), by(pln, "Revenue")
        cor_a, cor_p = by(act, "Cost of revenue"), by(pln, "Cost of revenue")
        opex_a = sum(by(act, f) for f in ("Research and development", "Sales and marketing",
                                          "General and administrative"))

        # ---------------------------------------------------- 1. the headline, tab by tab
        chk("Revenue on the P&L equals the Revenue tab",
            abs(num(val(P, "Revenue", 3)) - num(val(R, "Total revenue", 3))) < TOL)
        chk("Revenue on the P&L equals the ledger",
            abs(num(val(P, "Revenue", 3)) - rev_a) < TOL,
            f"{num(val(P, 'Revenue', 3)):,.2f} vs {rev_a:,.2f}")
        chk("Revenue benchmark is the plan of record",
            abs(num(val(P, "Revenue", 4)) - rev_p) < TOL,
            f"{num(val(P, 'Revenue', 4)):,.2f} vs {rev_p:,.2f}")
        chk("Cost of revenue on the P&L equals the Gross margin tab",
            abs(num(val(P, "Cost of revenue", 3)) - num(val(G, "Total cost of revenue", 3))) < TOL)
        chk("Cost of revenue equals the ledger",
            abs(num(val(P, "Cost of revenue", 3)) - cor_a) < TOL)
        chk("Gross profit is identical on both tabs",
            abs(num(val(P, "Gross profit", 3)) - num(val(G, "Gross profit", 3))) < TOL)
        chk("Revenue on the Gross margin tab is the Revenue tab's, not a restatement",
            abs(num(val(G, "Revenue", 3)) - num(val(R, "Total revenue", 3))) < TOL)

        # ---------------------------------------------------- 2. THE cross-tab check
        chk("Total operating expense by FSLI equals total by OWNER",
            abs(num(val(P, "Total operating expense", 3))
                - num(val(O, "Total operating expense", 5))) < TOL,
            f"{num(val(P, 'Total operating expense', 3)):,.2f}")
        for fsli in ("Research and development", "Sales and marketing",
                     "General and administrative"):
            chk(f"{fsli} agrees between the two cuts",
                abs(num(val(P, fsli, 3)) - num(val(O, fsli, 5))) < TOL,
                f"{num(val(P, fsli, 3)):,.2f} vs {num(val(O, fsli, 5)):,.2f}")
            chk(f"{fsli} equals the ledger",
                abs(num(val(P, fsli, 3)) - by(act, fsli)) < TOL)
        chk("Operating expense ties to the ledger in total",
            abs(num(val(P, "Total operating expense", 3)) - opex_a) < TOL)

        # ---------------------------------------------------- 3. balance sheet
        for col, label in ((3, "January"), (4, "December")):
            chk(f"Balance sheet balances, {label}",
                abs(num(val(B, "Check - must be nil", col))) < TOL,
                f"{num(val(B, 'Check - must be nil', col)):,.2f}")
        chk("Every balance sheet account is classified current or non-current",
            all(B.cell(row=r, column=7).value in ("Current", "Non-current", "Class", None)
                for r in range(1, B.max_row + 1)))
        chk("Quick ratio is never above the current ratio",
            num(val(B, "Quick ratio", 3)) <= num(val(B, "Current ratio", 3)) + 1e-9
            and num(val(B, "Quick ratio", 4)) <= num(val(B, "Current ratio", 4)) + 1e-9)
        chk("DSO including unbilled is never below DSO on trade",
            num(val(B, "DSO - including unbilled", 3)) >= num(val(B, "DSO - trade receivable", 3)))

        # ---------------------------------------------------- 4. cash
        for col, label in ((3, "January"), (4, "December")):
            chk(f"Cash walk reaches the ledger, {label}",
                abs(num(val(C, "Difference - must be nil", col))) < TOL,
                f"{num(val(C, 'Difference - must be nil', col)):,.2f}")
        chk("The walk's net result is the P&L's net result",
            abs(num(val(C, "Net result", 3)) - num(val(P, "NET RESULT", 3))) < TOL,
            f"{num(val(C, 'Net result', 3)):,.2f} vs {num(val(P, 'NET RESULT', 3)):,.2f}")
        chk("Closing cash per the walk equals the balance sheet's cash accounts",
            abs(num(val(C, "Closing cash, per the walk", 3))
                - num(val(C, "Closing cash, per the balance sheet", 3))) < TOL)
        chk("Runway is positive and finite", 0 < num(val(C, "Runway, months", 3)) < 240,
            f"{num(val(C, 'Runway, months', 3)):.1f} months")

        # ---------------------------------------------------- 5. ARR
        chk("The ARR waterfall foots to the ARR schedule",
            abs(num(val(R, "Difference", 3))) < TOL,
            f"{num(val(R, 'Difference', 3)):,.2f}")
        chk("Closing ARR is twelve times committed MRR",
            abs(num(val(R, "Closing ARR, 31 January 2026", 3))
                - num(val(R, "Committed MRR", 3)) * 12) < TOL)

        # ---------------------------------------------------- 6. EBITDA
        sbc = sum(act.get(a, 0.0) for a in ("6020", "7020", "8020"))
        chk("Adjusted EBITDA less EBITDA equals stock comp",
            abs((num(val(P, "ADJUSTED EBITDA", 3)) - num(val(P, "EBITDA", 3))) - sbc) < TOL,
            f"{num(val(P, 'ADJUSTED EBITDA', 3)) - num(val(P, 'EBITDA', 3)):,.2f} vs {sbc:,.2f}")
        chk("EBITDA is stated as well as adjusted EBITDA",
            find(P, "EBITDA") is not None and find(P, "ADJUSTED EBITDA") is not None)
        da = sum(act.get(a, 0.0) for a in ("5070", "8080"))
        chk("EBITDA less the operating result equals depreciation and amortization",
            abs((num(val(P, "EBITDA", 3)) - num(val(P, "OPERATING RESULT", 3))) - da) < TOL)

        # ---------------------------------------------------- 7. the hosted LBE
        for label, col in (("Total revenue", 3), ("OPERATING RESULT", 3),
                           ("OPERATING RESULT", 4), ("ADJUSTED EBITDA", 5)):
            a, b = val(LB, label, col, 1), val(LS, label, col, 1)
            chk(f"Hosted LBE equals the standalone artifact - {label} col {col}",
                abs(num(a) - num(b)) < TOL, f"{num(a):,.2f}")
        chk("The LBE identity holds: LBE = Fcst + Variance",
            abs(num(val(LS, "OPERATING RESULT", 5, 1))
                - (num(val(LS, "OPERATING RESULT", 3, 1))
                   + num(val(LS, "OPERATING RESULT", 4, 1)))) < TOL)
        # The approved changes are summed off the flux analysis workbook's own Q1 LBE effect
        # column, not from a file beside the builder. A check that depends on something
        # outside the instance cannot run where the instance actually lives.
        FX = load_workbook(recalc(FLUX, tmp), data_only=True)["Flux analysis"]
        approved = sum(c.value for row in FX.iter_rows(min_col=13, max_col=13)
                       for c in row if isinstance(c.value, (int, float)))
        approved /= 2      # the column carries each line once and each FSLI subtotal once
        chk("Every approved LBE change is in the forecast, and nothing else is",
            abs(num(val(LS, "OPERATING RESULT", 4, 1)) - approved) < TOL,
            f"{num(val(LS, 'OPERATING RESULT', 4, 1)):,.2f} vs {approved:,.2f} approved")
        chk("The LBE's plan column is the plan of record for the quarter",
            abs(num(val(LS, "Total revenue", 3, 1))
                - sum(float(r["amount_usd"]) for r
                      in read(F("02-budget", "FY2026", "plan_fy26_pnl_monthly.csv"))
                      if r["account"] and r["account"][0] == "4"
                      and r["period"] in ("2026-01", "2026-02", "2026-03"))) < TOL)

        # ---------------------------------------------------- 8. the pack's own rules
        blue = []
        formulas = []
        for ws in raw.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    fill = c.fill
                    if fill is not None and fill.fgColor is not None \
                            and fill.fgColor.rgb in ("FF0070C0", "FF0000FF", "FF00B0F0"):
                        blue.append(f"{ws.title}!{c.coordinate}")
                    if isinstance(c.value, str) and c.value.startswith("="):
                        formulas.append(c)
        chk("No cell is coloured as a typed input", not blue, f"{len(blue)} blue cells")
        chk("The pack is built from formulas", len(formulas) > 250, f"{len(formulas)} formulas")

        named = [("Balance sheet", "DSO - trade receivable", 3),
                 ("Balance sheet", "DPO", 3), ("Balance sheet", "Current ratio", 3),
                 ("Balance sheet", "Quick ratio", 3), ("Cash", "Runway, months", 3),
                 ("Cash", "Average monthly burn - months that consumed cash only", 3),
                 ("Gross margin", "Blended cost per page", 6),
                 ("P&L by FSLI", "Gross profit", 3),
                 ("P&L by FSLI", "Total operating expense", 3),
                 ("P&L by FSLI", "ADJUSTED EBITDA", 3),
                 ("Revenue and ARR", "Closing ARR, 31 January 2026", 3)]
        typed = []
        for sheet, label, col in named:
            ws = raw[sheet]
            r = find(ws, label, 2) or find(ws, label, 1)
            v = ws.cell(row=r, column=col).value if r else None
            if not (isinstance(v, str) and v.startswith("=")):
                typed.append(f"{sheet}!{label}")
        chk("Every ratio and every derived headline is a formula, not a number", not typed,
            f"{typed}")

        # ---------------------------------------------------- 9. no plugs
        chk("The revenue bridge to plan has no residual",
            abs(num(val(P, "Revenue", 5)) - (rev_a - rev_p)) < TOL)
        # Excel error values only. A bare "#" is a legitimate label - it is the header of the
        # open-items table - and matching on the character rather than the error flagged it.
        ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!")
        errs = [f"{ws.title}!{c.coordinate}={c.value}" for ws in wb.worksheets
                for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.strip() in ERRORS]
        chk("Nothing on the pack recalculated to an error", not errs, f"{errs[:4]}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

    bad = [r for r in results if not r[0]]
    print(f"\n{len(results) - len(bad)} of {len(results)} checks pass")
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
